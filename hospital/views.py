"""
hospital/views.py
Django views replacing CSV operations with ORM queries.
"""
import json
import logging
from collections import Counter
from datetime import datetime, date, timedelta
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.core.cache import cache

from .models import Department, Doctor, Nurse, Patient, Visit, HospitalConfiguration
from .forms import HospitalConfigurationForm
from ml.predict_load import (
    predict_department_load_from_df,
    predict_department_load_over_time_from_df,
    analyze_staffing_from_df,
    predict_medicine_demand_from_df,
    analyze_peak_traffic_from_df
)

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------
#  Helpers
# ----------------------------------------------------------------------

def calculate_age(dob_obj):
    """Calculate age from date object."""
    if not dob_obj:
        return ""
    if isinstance(dob_obj, str):
        try:
            dob_obj = datetime.strptime(dob_obj, "%Y-%m-%d").date()
        except:
            return ""
            
    today = date.today()
    age = today.year - dob_obj.year - ((today.month, today.day) < (dob_obj.month, dob_obj.day))
    return age


# ----------------------------------------------------------------------
#  Hospital Settings
# ----------------------------------------------------------------------
@login_required
@user_passes_test(lambda u: u.is_superuser)
def edit_hospital_config(request):
    config = HospitalConfiguration.get_solo()
    if request.method == 'POST':
        form = HospitalConfigurationForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Hospital settings updated successfully.")
            return redirect('hospital-settings')
    else:
        form = HospitalConfigurationForm(instance=config)
    
    return render(request, "hospital/edit_config.html", {"form": form, "config": config})


# ----------------------------------------------------------------------
#  Auth Views (Handled by accounts app now)
# ----------------------------------------------------------------------




# ----------------------------------------------------------------------
#  Hub & Staff Views
# ----------------------------------------------------------------------
@login_required
def hub_view(request):
    """
    Dispatcher view: Redirects users to their specific dashboard based on role.
    """
    user = request.user
    if user.is_superuser:
        return redirect('admin_dashboard')
    
    # Check Groups
    if user.groups.filter(name='Receptionist').exists() or user.username == 'receptionist':
        return redirect('reception_dashboard')
    elif user.groups.filter(name='Doctor').exists():
        return redirect('doctor_dashboard')
    elif user.groups.filter(name='Pharmacy').exists() or user.username == 'pharmacy':
        return redirect('pharmacy_inventory')
        
    # Check Models directly (fallback for doctors not in Group)
    if Doctor.objects.filter(staff_id=user.username).exists():
        return redirect('doctor_dashboard')
    elif Nurse.objects.filter(staff_id=user.username).exists():
        # Nurses might share Reception dashboard or have their own view? defaulting to Reception for now or Hub
        return redirect('reception_dashboard') 
        
    # Fallback for generic staff
    return render(request, "hub.html")

@login_required
def admin_dashboard(request):
    return render(request, "dashboards/admin.html")

@login_required
def reception_dashboard(request):
    # Similar context to original index.html but for reception
    context = {
        "doctors_by_dept": get_doc_map(),
        "departments": list(Department.objects.values_list('name', flat=True).order_by('name')),
        "is_reception": True
    }
    return render(request, "dashboards/reception.html", context)

@login_required
def doctor_dashboard(request):
    # Filter visits for this doctor if possible
    # We try to match the logged-in user's username to the Doctor's staff_id
    try:
        doctor = Doctor.objects.get(staff_id=request.user.username)
        my_visits = Visit.objects.filter(doctor=doctor).select_related('patient').order_by('-visit_date')
    except Doctor.DoesNotExist:
        # If user is in Doctor group but not in Doctor table (e.g. mismatch), show empty
        my_visits = []
        doctor = None

    context = {
        "my_visits": my_visits,
        "doctor": doctor
    }
    return render(request, "dashboards/doctor.html", context)

@login_required
def doctor_analytics(request):
    try:
        doctor = Doctor.objects.get(staff_id=request.user.username)
    except Doctor.DoesNotExist:
        return redirect('hub') # Logic fail fallback

    # Use visits_qs to avoid repeated queries
    visits_qs = Visit.objects.filter(doctor=doctor)
    
    # 1. Basic Stats (Aggregated in DB)
    total_visits = Visit.objects.filter(doctor=doctor).count()
    today_visits = Visit.objects.filter(doctor=doctor, visit_date=date.today()).count()
    unique_patients = Visit.objects.filter(doctor=doctor).values('patient').distinct().count()
    
    # 2. Case Type Distribution
    case_dist = Visit.objects.filter(doctor=doctor).values('case_type').annotate(count=Count('id'))
    case_labels = [c['case_type'] for c in case_dist]
    case_data = [c['count'] for c in case_dist]
    
    # 3. Daily Activity (Last 7 Days)
    seven_days_ago = date.today() - timedelta(days=6)
    daily_stats = Visit.objects.filter(doctor=doctor, visit_date__gte=seven_days_ago)\
                           .values('visit_date')\
                           .annotate(count=Count('id'))\
                           .order_by('visit_date')
                           
    daily_dict = {d['visit_date']: d['count'] for d in daily_stats}
    trend_labels = []
    trend_data = []
    current = seven_days_ago
    while current <= date.today():
        trend_labels.append(current.strftime("%d %b"))
        trend_data.append(daily_dict.get(current, 0))
        current += timedelta(days=1)
        
    # 4. Department Prediction (Future Load) - Cached & Limited
    cache_key = f'doctor_analytics_pred_{doctor.department.name}'
    pred_results = cache.get(cache_key)
    
    if not pred_results:
        # Only fetch recent 10k visits for this department to keep it fast
        dept_visits = Visit.objects.filter(department=doctor.department)\
                                  .order_by('-visit_date')[:10000]\
                                  .values('department__name', 'visit_date')
        if dept_visits:
             df = pd.DataFrame(list(dept_visits))
             df.rename(columns={'department__name': 'department'}, inplace=True)
             preds = predict_department_load_over_time_from_df(df, months=3)
             if doctor.department.name in preds:
                 pred_results = preds[doctor.department.name]
        cache.set(cache_key, pred_results, 3600) # Cache for 1 hour
    
    pred_data = pred_results or []
    
    # X-axis labels for the prediction chart (next 3 months)
    today_val = date.today()
    pred_labels = []
    for i in range(1, 4):
        future = today_val + timedelta(days=30*i) 
        pred_labels.append(future.strftime("%b %Y"))
    
    context = {
        "doctor": doctor,
        "total_visits": total_visits,
        "today_visits": today_visits,
        "unique_patients": unique_patients,
        "case_labels": json.dumps(case_labels),
        "case_data": json.dumps(case_data),
        "trend_labels": json.dumps(trend_labels),
        "trend_data": json.dumps(trend_data),
        "pred_labels": json.dumps(pred_labels),
        "pred_data": json.dumps(pred_data)
    }
    return render(request, "dashboards/doctor_analytics.html", context)

@login_required
def doctor_specialized_insights(request):
    try:
        doctor = Doctor.objects.get(staff_id=request.user.username)
    except Doctor.DoesNotExist:
        return redirect('hub')
        
    # 1. Fetch recent visits specific to this doctor
    recent_visits = Visit.objects.filter(doctor=doctor).select_related('patient', 'department').order_by('-visit_date')[:50]
    
    # Format visits for the template table
    formatted_visits = []
    for v in recent_visits:
        formatted_visits.append({
            "id": v.patient.patient_id,
            "name": v.patient.patient_name,
            "dept": v.department.name,
            "case_type": v.case_type,
            "date": str(v.visit_date),
            "medicines": v.medicines,
            "symptoms": v.symptoms
        })

    # Stats for the top section
    visits_qs = Visit.objects.filter(doctor=doctor)
    total_visits = visits_qs.count()
    unique_patients = visits_qs.values('patient').distinct().count()
    
    # 2. Personalized AI Predictions over time
    pred_data = {}
    doctor_visits = visits_qs.values('doctor__doctor_name', 'visit_date')
    if doctor_visits:
        df = pd.DataFrame(list(doctor_visits))
        # Rename column so the ML function treats the doctor as the "department" 
        df.rename(columns={'doctor__doctor_name': 'department'}, inplace=True)
        # Predict for next 3 months
        preds = predict_department_load_over_time_from_df(df, months=3)
        if doctor.doctor_name in preds:
            pred_data = preds[doctor.doctor_name]
            
    # X-axis labels for the prediction chart (next 3 months)
    today_val = date.today()
    pred_labels = []
    for i in range(1, 4):
        future = today_val + timedelta(days=30*i) 
        pred_labels.append(future.strftime("%b %Y"))

    context = {
        "doctor": doctor,
        "total_visits": total_visits,
        "unique_patients": unique_patients,
        "recent_visits": formatted_visits,
        "pred_labels": json.dumps(pred_labels),
        "pred_data": json.dumps(pred_data) if isinstance(pred_data, list) else json.dumps([])
    }
    return render(request, "dashboards/doctor_specialized.html", context)

@login_required
def staff_view(request):
    doctors = Doctor.objects.select_related('department').all()
    nurses = Nurse.objects.select_related('department').all()
    departments = Department.objects.values_list('name', flat=True).order_by('name')

    doc_list = []
    for d in doctors:
        doc_list.append({
            "staff_id": d.staff_id,
            "department": d.department.name,
            "name": d.doctor_name,
            "specialization": d.specialization,
            "gender": d.gender,
            "qualifications": d.qualifications,
            "email": d.email,
            "phone": d.phone
        })

    nurse_list = []
    for n in nurses:
        nurse_list.append({
            "staff_id": n.staff_id,
            "department": n.department.name,
            "name": n.nurse_name,
            "gender": n.gender,
            "qualifications": n.qualifications,
            "email": n.email,
            "phone": n.phone
        })

    return render(request, "staff.html", {
        "doctors": doc_list, 
        "nurses": nurse_list,
        "departments": list(departments)
    })

@login_required
def add_staff(request):
    if request.method == "POST":
        role = request.POST.get("role") # "doctor" or "nurse"
        name = request.POST.get("name")
        dept_name = request.POST.get("department")
        gender = request.POST.get("gender", "")
        qual = request.POST.get("qualifications", "")
        email = request.POST.get("email", "")
        phone = request.POST.get("phone", "")
        spec = request.POST.get("specialization", "")

        department, _ = Department.objects.get_or_create(name=dept_name)

        # Generate ID
        prefix = "DOC" if role == "doctor" else "NUR"
        model_cls = Doctor if role == "doctor" else Nurse
        
        # Simple ID generation logic: find max numeric part
        last_obj = model_cls.objects.filter(staff_id__startswith=prefix).order_by('-staff_id').first()
        if last_obj:
            try:
                # Extract numeric part assuming format XX000 or similar
                # If ID is "DOC100", starts with DOC (3 chars)
                last_num = int(last_obj.staff_id[3:])
                new_num = last_num + 1
            except:
                new_num = 1
        else:
            new_num = 1
            
        staff_id = f"{prefix}{new_num:03d}"

        if role == "doctor":
            Doctor.objects.create(
                staff_id=staff_id,
                department=department,
                doctor_name=name,
                gender=gender,
                qualifications=qual,
                email=email,
                phone=phone,
                specialization=spec,
                # Defaults
                shift_start="09:00",
                shift_end="17:00",
                work_days="Mon,Tue,Wed,Thu,Fri",
                shift_type="Morning",
                slot_duration=15,
                max_patients=25
            )
        else:
            Nurse.objects.create(
                staff_id=staff_id,
                department=department,
                nurse_name=name,
                gender=gender,
                qualifications=qual,
                email=email,
                phone=phone,
                # Defaults
                shift_start="09:00",
                shift_end="17:00",
                work_days="Mon,Tue,Wed,Thu,Fri",
                shift_type="Morning"
            )
            
        messages.success(request, f"Successfully created {role.capitalize()}: {name} ({staff_id})")
        return redirect('staff_dashboard')
    
    # GET Request
    departments = Department.objects.values_list('name', flat=True).order_by('name')
    return render(request, "add_staff.html", {"departments": departments})

@csrf_exempt
@login_required
def edit_staff(request):
    role = request.POST.get("role")
    staff_id = request.POST.get("staff_id")
    dept_name = request.POST.get("department")
    
    try:
        department, _ = Department.objects.get_or_create(name=dept_name)
        
        if role == "doctor":
            staff = Doctor.objects.get(staff_id=staff_id)
            staff.specialization = request.POST.get("specialization", staff.specialization)
            staff.doctor_name = request.POST.get("name", staff.doctor_name)
        else:
            staff = Nurse.objects.get(staff_id=staff_id)
            staff.nurse_name = request.POST.get("name", staff.nurse_name)
            
        staff.department = department
        staff.gender = request.POST.get("gender", staff.gender)
        staff.qualifications = request.POST.get("qualifications", staff.qualifications)
        staff.email = request.POST.get("email", staff.email)
        staff.phone = request.POST.get("phone", staff.phone)
        staff.save()
        
        return JsonResponse({"status": "ok"})
    except ObjectDoesNotExist:
        return JsonResponse({"status": "error", "message": "Staff not found"})

@csrf_exempt
@login_required
def delete_staff(request):
    role = request.POST.get("role")
    staff_id = request.POST.get("staff_id")
    
    try:
        if role == "doctor":
            Doctor.objects.filter(staff_id=staff_id).delete()
        else:
            Nurse.objects.filter(staff_id=staff_id).delete()
        return JsonResponse({"status": "deleted"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@login_required
def department_dashboard(request):
    departments = Department.objects.values_list('name', flat=True).order_by('name')
    return render(request, "departments.html", {"departments": list(departments)})

@csrf_exempt
@login_required
def add_department(request):
    name = request.POST.get("name")
    if not name:
        return JsonResponse({"status": "error", "message": "Name is required"})
    
    if Department.objects.filter(name=name).exists():
        return JsonResponse({"status": "error", "message": "Department already exists"})
    
    Department.objects.create(name=name)
    return JsonResponse({"status": "ok"})

@csrf_exempt
@login_required
def delete_department(request):
    name = request.POST.get("name")
    try:
        Department.objects.filter(name=name).delete()
        return JsonResponse({"status": "deleted"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@csrf_exempt
@login_required
def edit_department(request):
    old_name = request.POST.get("old_name")
    new_name = request.POST.get("new_name")
    
    if not old_name or not new_name:
        return JsonResponse({"status": "error", "message": "Required fields missing"})

    try:
        with transaction.atomic():
            dept = Department.objects.get(name=old_name)
            # Create new department
            new_dept, _ = Department.objects.get_or_create(name=new_name)
            
            # Move all related objects to new department
            Doctor.objects.filter(department=dept).update(department=new_dept)
            Nurse.objects.filter(department=dept).update(department=new_dept)
            Visit.objects.filter(department=dept).update(department=new_dept)
            
            # Delete old department
            dept.delete()
            
        return JsonResponse({"status": "ok"})
    except ObjectDoesNotExist:
         return JsonResponse({"status": "error", "message": "Department not found"})
    except Exception as e:
         return JsonResponse({"status": "error", "message": str(e)})

# ----------------------------------------------------------------------
#  Reports View
# ----------------------------------------------------------------------
def get_doc_map():
    # Build dictionary { "DeptName": [ {staff_id, name}, ... ] }
    doc_map = {}
    doctors = Doctor.objects.select_related('department').all()
    for d in doctors:
        dept = d.department.name
        doc_obj = {"staff_id": d.staff_id, "name": d.doctor_name}
        doc_map.setdefault(dept, []).append(doc_obj)
    return doc_map

@login_required
def reports_view(request):
    departments = Department.objects.values_list('name', flat=True).order_by('name')
    doc_map = json.dumps(get_doc_map())
    return render(request, "reports.html", {
        "departments": list(departments), 
        "doctor_map_json": doc_map
    })

def filter_visits(request):
    """Common filtering logic for visits"""
    filter_type = request.GET.get("type") # "date" or "month"
    value = request.GET.get("value")      
    dept_filter = request.GET.get("department")
    doc_filter = request.GET.get("doctor")

    visits = Visit.objects.select_related('patient', 'doctor', 'department').all()

    if dept_filter:
        visits = visits.filter(department__name=dept_filter)
    
    if doc_filter:
        visits = visits.filter(doctor__doctor_name=doc_filter) # Assuming filtering by name string from frontend

    if filter_type == "date" and value:
        visits = visits.filter(visit_date=value)
    elif filter_type == "month" and value:
        # value is YYYY-MM
        visits = visits.filter(visit_date__startswith=value)
    
    return visits.order_by('-visit_date')

@login_required
def api_reports(request):
    visits = filter_visits(request)
    data = []
    for v in visits:
        data.append({
            "id": v.patient.patient_id,
            "name": v.patient.patient_name,
            "dept": v.department.name,
            "doc": v.doctor.doctor_name,
            "case_type": v.case_type,
            "date": str(v.visit_date)
        })
    return JsonResponse({"visits": data})

@login_required
def export_reports(request):
    visits = filter_visits(request)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"visit_report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visit Report"

    # Header
    headers = ["Date", "Patient ID", "Patient Name", "Department", "Doctor", "Case Type", "Medicines", "Symptoms"]
    ws.append(headers)
    
    # Style header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)

    for v in visits:
        ws.append([
            v.visit_date,
            v.patient.patient_id,
            v.patient.patient_name,
            v.department.name,
            v.doctor.doctor_name,
            v.case_type,
            v.medicines,
            v.symptoms
        ])
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

# ----------------------------------------------------------------------
#  Patient Dashboard
# ----------------------------------------------------------------------
@login_required
def patient_dashboard(request):
    context = {
        "doctors_by_dept": get_doc_map(),
        "departments": list(Department.objects.values_list('name', flat=True).order_by('name'))
    }
    return render(request, "index.html", context)

@login_required
def load_patients(request):
    """API to return all patients as JSON for frontend lookup. Optimized for large datasets."""
    # Use .values() to fetch only what we need, which is MUCH faster than model instances
    patients = Patient.objects.all().order_by('patient_id').values(
        'patient_id', 'patient_name', 'age', 'dob', 'gender', 'location', 'assigned_slot'
    )
    
    data = {}
    for p in patients:
        data[p['patient_id']] = {
            "name": p['patient_name'],
            "age": p['age'],
            "dob": p['dob'].strftime("%Y-%m-%d") if p['dob'] else "",
            "gender": p['gender'],
            "location": p['location'],
            "assigned_slot": p['assigned_slot']
        }
    return JsonResponse(data)

def get_next_slot_logic(doctor_name):
    """
    Calculates next available slot for a doctor.
    """
    doctor = Doctor.objects.filter(doctor_name=doctor_name).first()
    if not doctor:
        return "Unknown Schedule"

    start_time = doctor.shift_start
    # Naive date: today
    today = date.today()
    
    # Count patients assigned to this doctor TODAY (or loosely in the system if that was legacy logic)
    # The legacy logic counted ALL patients in patients.csv assigned to this doctor
    # To mimic that specifically for the "add patient" flow:
    patient_count = Patient.objects.filter(doctor=doctor).count()
    
    if patient_count >= doctor.max_patients:
        return "Full"

    # Calculate time
    # Convert time objects to minutes from midnight
    start_mins = start_time.hour * 60 + start_time.minute
    duration = doctor.slot_duration
    
    slot_start_mins = start_mins + (patient_count * duration)
    slot_end_mins = slot_start_mins + duration
    
    def fmt(mins):
        th = (mins // 60) % 24
        tm = mins % 60
        return f"{th:02}:{tm:02}"
        
    return f"{fmt(slot_start_mins)} - {fmt(slot_end_mins)}"

def get_next_slot(doctor_name):
    # This might be called from view logic or template tags if needed
    return get_next_slot_logic(doctor_name)

@csrf_exempt
def add_patient(request):
    pid = request.POST.get("id")
    name = request.POST.get("name")
    doctor_name = request.POST.get("doctor", "")
    dob_str = request.POST.get("dob", "")
    gender = request.POST.get("gender", "")
    location = request.POST.get("location", "")
    
    if Patient.objects.filter(patient_id=pid).exists():
         return JsonResponse({"status": "error", "message": "Patient ID already exists"})

    dob = None
    age = request.POST.get("age", 0)
    if dob_str:
        try:
            dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
            age = calculate_age(dob)
        except:
            pass
            
    doctor = None
    assigned_slot = ""
    if doctor_name:
        doctor = Doctor.objects.filter(doctor_name=doctor_name).first()
        if doctor:
            assigned_slot = get_next_slot_logic(doctor_name)

    Patient.objects.create(
        patient_id=pid,
        patient_name=name,
        age=int(age) if age else 0,
        dob=dob,
        gender=gender,
        location=location,
        doctor=doctor,
        assigned_slot=assigned_slot
    )
    return JsonResponse({"status": "ok"})

@csrf_exempt
def delete_patient(request):
    pid = request.POST.get("patient_id")
    try:
        Patient.objects.filter(patient_id=pid).delete()
        # Visits are cascaded if on_delete=CASCADE in model
        return JsonResponse({"status": "deleted"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

@csrf_exempt
def edit_patient(request):
    pid = request.POST.get("patient_id")
    try:
        patient = Patient.objects.get(patient_id=pid)
        patient.patient_name = request.POST.get("new_name", patient.patient_name)
        
        new_dob = request.POST.get("new_dob")
        if new_dob:
            patient.dob = datetime.strptime(new_dob, "%Y-%m-%d").date()
            patient.age = calculate_age(patient.dob)
        elif request.POST.get("new_age"):
            patient.age = int(request.POST.get("new_age"))
            
        patient.gender = request.POST.get("new_gender", patient.gender)
        patient.location = request.POST.get("new_location", patient.location)
        patient.save()
        return JsonResponse({"status": "ok"})
    except Patient.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Patient not found"})

# ----------------------------------------------------------------------
#  Visit CRUD
# ----------------------------------------------------------------------
@csrf_exempt
def add_visit(request):
    pid = request.POST.get("patient_id")
    dept_name = request.POST.get("department")
    doc_name = request.POST.get("doctor")
    case_type = request.POST.get("case_type")
    
    try:
        patient = Patient.objects.get(patient_id=pid)
        department, _ = Department.objects.get_or_create(name=dept_name)
        doctor = Doctor.objects.filter(doctor_name=doc_name).first()
        if not doctor:
            raise ObjectDoesNotExist(f"Doctor '{doc_name}' not found.")
        
        slot = get_next_slot_logic(doc_name)
        
        raw_medicines = request.POST.get("medicines", "")
        # The frontend will now send JSON string of medicines: [{"name": "Aspirin", "qty": 10, "dosage": "1-0-1"}]
        med_summary = []
        parsed_meds = []
        
        try:
            # Try to parse as JSON first (new format)
            if raw_medicines.startswith("["):
                parsed_meds = json.loads(raw_medicines)
                for m in parsed_meds:
                    summary = f"{m.get('name', '')} x{m.get('qty', 1)}"
                    if m.get('dosage'):
                        summary += f" ({m.get('dosage')})"
                    med_summary.append(summary)
            else:
                # Fallback to old format (comma-separated string)
                if raw_medicines:
                    med_summary = [m.strip() for m in raw_medicines.split(",") if m.strip()]
        except json.JSONDecodeError:
            if raw_medicines:
                med_summary = [m.strip() for m in raw_medicines.split(",") if m.strip()]
                
        visit_date_str = request.POST.get("visit_date")
        if visit_date_str:
            visit_date_obj = datetime.strptime(visit_date_str, "%Y-%m-%d").date()
            if visit_date_obj > date.today():
                return JsonResponse({"status": "error", "message": "Visit date cannot be in the future."})

        visit = Visit.objects.create(
            patient=patient,
            department=department,
            doctor=doctor,
            case_type=case_type,
            visit_date=visit_date_str,
            medicines=", ".join(med_summary),
            symptoms=request.POST.get("symptoms", ""),
            slot=slot
        )
        
        # Create PrescriptionLine objects for the new format
        from pharmacy.models import Medicine, PrescriptionLine
        for m in parsed_meds:
            med_name = m.get('name', '').strip()
            qty = int(m.get('qty', 1))
            dosage = m.get('dosage', '').strip()
            
            try:
                # Find the medicine in DB. We only create prescription line if it exists in Pharmacy.
                med_obj = Medicine.objects.get(name__iexact=med_name)
                PrescriptionLine.objects.create(
                    visit=visit,
                    medicine=med_obj,
                    quantity=qty,
                    dosage_instructions=dosage
                )
            except Medicine.DoesNotExist:
                logger.warning(f"Add Visit: Medicine '{med_name}' not found in DB.")
                
        return JsonResponse({"status": "ok"})
    except ObjectDoesNotExist as e:
         return JsonResponse({"status": "error", "message": str(e)})
    except Exception as e:
         return JsonResponse({"status": "error", "message": str(e)})

@csrf_exempt
def delete_visit(request):
    # Deleting by composite keys is tricky without an ID.
    # The original tracked row content.
    # We will try to find a match.
    pid = request.POST.get("patient_id")
    date = request.POST.get("visit_date")
    # Finding purely by these two might be risky if multiple visits same day, 
    # but that's the best we can do with current frontend.
    
    Visit.objects.filter(
        patient__patient_id=pid,
        visit_date=date,
        case_type=request.POST.get("case_type")
    ).delete()
    
    # Ideally should delete only one if multiple match
    return JsonResponse({"status": "deleted"})

@csrf_exempt
def edit_visit(request):
    # This is also tricky without Visit ID.
    # We use "old" values to find the record.
    pid = request.POST.get("patient_id")
    o_date = request.POST.get("old_visit_date")
    o_case = request.POST.get("old_case_type")

    try:
        # Get the specific visit
        visit = Visit.objects.filter(
            patient__patient_id=pid,
            visit_date=o_date,
            case_type=o_case
        ).first()
        
        if visit:
            visit_date_str = request.POST.get("visit_date")
            if visit_date_str:
                visit_date_obj = datetime.strptime(visit_date_str, "%Y-%m-%d").date()
                if visit_date_obj > date.today():
                    return JsonResponse({"status": "error", "message": "Visit date cannot be in the future."})
            visit.visit_date = visit_date_str

            visit.case_type = request.POST.get("case_type")
            visit.symptoms = request.POST.get("symptoms", "")
            
            raw_medicines = request.POST.get("medicines", "")
            med_summary = []
            parsed_meds = []
            
            try:
                if raw_medicines.startswith("["):
                    parsed_meds = json.loads(raw_medicines)
                    for m in parsed_meds:
                        summary = f"{m.get('name', '')} x{m.get('qty', 1)}"
                        if m.get('dosage'):
                            summary += f" ({m.get('dosage')})"
                        med_summary.append(summary)
                else:
                    if raw_medicines:
                        med_summary = [m.strip() for m in raw_medicines.split(",") if m.strip()]
            except json.JSONDecodeError:
                if raw_medicines:
                    med_summary = [m.strip() for m in raw_medicines.split(",") if m.strip()]
                    
            visit.medicines = ", ".join(med_summary)
            
            new_dept = request.POST.get("department")
            if new_dept:
                dept, _ = Department.objects.get_or_create(name=new_dept)
                visit.department = dept
                
            new_doc = request.POST.get("doctor")
            if new_doc:
                doc = Doctor.objects.filter(doctor_name=new_doc).first()
                if doc:
                    visit.doctor = doc
            
            visit.save()
            
            # Recreate PrescriptionLines from parsed data
            from pharmacy.models import Medicine, PrescriptionLine
            PrescriptionLine.objects.filter(visit=visit).delete()
            for m in parsed_meds:
                med_name = m.get('name', '').strip()
                qty = int(m.get('qty', 1))
                dosage = m.get('dosage', '').strip()
                try:
                    med_obj = Medicine.objects.get(name__iexact=med_name)
                    PrescriptionLine.objects.create(
                        visit=visit,
                        medicine=med_obj,
                        quantity=qty,
                        dosage_instructions=dosage
                    )
                except Medicine.DoesNotExist:
                    pass
                    
            return JsonResponse({"status": "ok"})
        else:
            return JsonResponse({"status": "error", "message": "Visit not found"})

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})

def load_visits(request, patient_id):
    visits = Visit.objects.filter(patient__patient_id=patient_id).select_related('department', 'doctor').order_by('-visit_date')
    
    # Pre-fetch prescriptions to send structured data to the frontend for editing
    from pharmacy.models import PrescriptionLine
    visit_ids = [v.id for v in visits]
    prescriptions = PrescriptionLine.objects.filter(visit_id__in=visit_ids).select_related('medicine')
    
    rx_map = {}
    for p in prescriptions:
        rx_map.setdefault(p.visit_id, []).append({
            "name": p.medicine.name,
            "qty": p.quantity,
            "dosage": p.dosage_instructions
        })
        
    data = []
    for v in visits:
        # If legacy visit, try to create basic JSON from text
        rx_data = rx_map.get(v.id, [])
        if not rx_data and v.medicines:
            med_names = [m.strip() for m in v.medicines.split(',') if m.strip()]
            for mn in med_names:
                import re
                qty = 1
                dosage = ""
                name = mn
                
                # parse (dosage)
                dos_match = re.search(r'\((.*?)\)', name)
                if dos_match:
                    dosage = dos_match.group(1).strip()
                    name = name.replace(dos_match.group(0), '')
                    
                # parse xQty or ×Qty
                qty_matches = re.findall(r'[x×X](\d+)', name)
                if qty_matches:
                    qty = int(qty_matches[-1])
                    name = re.sub(r'[x×X]\d+', '', name)
                    
                name = name.strip()
                rx_data.append({"name": name, "qty": qty, "dosage": dosage})
                
        rx_json = json.dumps(rx_data)
        
        # Frontend expects raw list of strings matching CSV format:
        # patient_id,department,doctor,case_type,visit_date,medicines,symptoms,slot,rx_json
        data.append([
            v.patient.patient_id,
            v.department.name,
            v.doctor.doctor_name,
            v.case_type,
            str(v.visit_date),
            v.medicines,
            v.symptoms,
            v.slot or "",
            rx_json
        ])
    return JsonResponse(data, safe=False)

def export_visits(request):
    pid = request.GET.get("patient_id")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"visits_{pid}.xlsx" if pid else "hospital_visits.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visits"
    
    headers = ["patient_id", "department", "doctor", "case_type", "visit_date", "medicines", "symptoms", "slot"]
    ws.append(headers)
    
    # Style header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)

    qs = Visit.objects.select_related('patient', 'department', 'doctor')
    if pid:
        qs = qs.filter(patient__patient_id=pid)
        
    for v in qs:
        ws.append([
            v.patient.patient_id,
            v.department.name,
            v.doctor.doctor_name,
            v.case_type,
            v.visit_date,
            v.medicines,
            v.symptoms,
            v.slot
        ])
    
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

# ----------------------------------------------------------------------
#  ML Helpers
# ----------------------------------------------------------------------
@login_required
def show_prediction(request):
    return render(request, "prediction.html")

def predict_load(request):
    cache_key = 'hospital_prediction_load'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)

    # Limit data for prediction to recent 100k for performance
    visits = Visit.objects.all().order_by('-visit_date')[:250000].values('department__name', 'visit_date')
    if not visits:
        return JsonResponse({})
        
    df = pd.DataFrame(list(visits))
    df.rename(columns={'department__name': 'department'}, inplace=True)
    
    data = predict_department_load_from_df(df)
    # Cache for 1 hour
    cache.set(cache_key, data, 3600)
    return JsonResponse(data)

def predict_load_over_time(request):
    months = int(request.GET.get("months", 3))
    cache_key = f'hospital_prediction_load_over_time_{months}'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)

    visits = Visit.objects.all().order_by('-visit_date')[:250000].values('department__name', 'visit_date')
    if not visits:
        return JsonResponse({})
        
    df = pd.DataFrame(list(visits))
    df.rename(columns={'department__name': 'department'}, inplace=True)
    
    data = predict_department_load_over_time_from_df(df, months=months)
    # Cache for 1 hour
    cache.set(cache_key, data, 3600)
    return JsonResponse(data)

def get_suggestions(request):
    symptoms = set()
    medicines = set()
    
    import re
    # 1. Fetch historical text-based data
    visits = Visit.objects.all().values('symptoms', 'medicines')
    for v in visits:
        if v['symptoms']:
            for s in v['symptoms'].split(','):
                if s.strip(): symptoms.add(s.strip())
        if v['medicines']:
            for m in v['medicines'].split(','):
                name = m.strip()
                if not name: continue
                # Strip (dosage)
                dos_match = re.search(r'\((.*?)\)', name)
                if dos_match:
                    name = name.replace(dos_match.group(0), '')
                # Strip xQty
                qty_match = re.search(r'[x×X](\d+)', name)
                if qty_match:
                    name = name.replace(qty_match.group(0), '')
                name = name.strip()
                if name:
                    medicines.add(name)
    
    # 2. Fetch from Pharmacy Inventory (Real Data)
    from pharmacy.models import Medicine
    inventory = Medicine.objects.values_list('name', flat=True)
    for name in inventory:
        medicines.add(name)

    return JsonResponse({
        "symptoms": sorted(list(symptoms)),
        "medicines": sorted(list(medicines))
    })

def analytics_data(request):
    cache_key = 'hospital_analytics_data'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)

    # Heavy analytics - efficiently done with DB aggregation
    from django.db.models import Count
    from collections import Counter
    
    # 1. Direct DB aggregations (Fast)
    doctor_load_qs = Visit.objects.values('doctor__staff_id', 'doctor__doctor_name').annotate(count=Count('id')).order_by('-count')
    dept_load_qs = Visit.objects.values('department__name').annotate(count=Count('id')).order_by('-count')
    cases_type_qs = Visit.objects.values('case_type').annotate(count=Count('id')).order_by('-count')
    
    doctor_load = {f"{item['doctor__doctor_name']} ({item['doctor__staff_id']})": item['count'] for item in doctor_load_qs}
    dept_load = {item['department__name']: item['count'] for item in dept_load_qs}
    cases_type = {item['case_type']: item['count'] for item in cases_type_qs}
    
    # 2. Text-based processing (Sampled for performance)
    # Only use recent 10k visits for symptom/medicine analysis to be fast (statistically significant sample)
    recent_visits = Visit.objects.all().order_by('-visit_date')[:10000].values('symptoms', 'medicines')
    
    symptom_dist = Counter()
    medicine_usage = Counter()
    
    import re
    def clean_med(name):
        name = name.strip()
        if not name: return None
        # Strip (dosage) and xQty
        name = re.sub(r'\(.*?\)', '', name)
        name = re.sub(r'[x×X]\d+', '', name)
        return name.strip()

    for v in recent_visits:
        if v['symptoms']:
            for s in v['symptoms'].split(','):
                if s.strip(): symptom_dist[s.strip()] += 1
        if v['medicines']:
            for m in v['medicines'].split(','):
                if m.strip(): medicine_usage[m.strip()] += 1
                
    # 3. Monthly Trend (Native DB Aggregation across all records for an accurate graph)
    from django.db.models.functions import TruncMonth
    trend_qs = Visit.objects.annotate(month=TruncMonth('visit_date')).values('month').annotate(c=Count('id')).order_by('month')
    monthly_trend = {item['month'].strftime('%Y-%m'): item['c'] for item in trend_qs if item['month']}

    data = {
        "doctor_load": doctor_load,
        "dept_load": dept_load,
        "cases_type": cases_type,
        "symptom_dist": dict(symptom_dist.most_common(20)),
        "medicine_usage": dict(medicine_usage.most_common(20)),
        "monthly_trend": dict(sorted(monthly_trend.items())),
        "metadata": {
            "total_visits": sum(dept_load.values()),
            "top_department": dept_load_qs[0]['department__name'] if dept_load_qs else "N/A",
            "total_doctors": Doctor.objects.count(),
        }
    }
    # Cache for 1 hour
    cache.set(cache_key, data, 3600)
    return JsonResponse(data)

def resource_planning(request):
    cache_key = 'hospital_resource_planning'
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)

    # Optimization: Use most recent 100k visits for planning analysis
    visits_qs = Visit.objects.all().order_by('-visit_date')[:250000].values('department__name', 'visit_date', 'medicines', 'symptoms')
    visits_df = pd.DataFrame(list(visits_qs))
    if not visits_df.empty:
        visits_df.rename(columns={'department__name': 'department'}, inplace=True)
    
    doctors_qs = Doctor.objects.all().values('department__name')
    doctors_df = pd.DataFrame(list(doctors_qs))
    if not doctors_df.empty:
        doctors_df.rename(columns={'department__name': 'department'}, inplace=True)
        
    nurses_qs = Nurse.objects.all().values('department__name')
    nurses_df = pd.DataFrame(list(nurses_qs))
    if not nurses_df.empty:
        nurses_df.rename(columns={'department__name': 'department'}, inplace=True)
        
    staffing_data = analyze_staffing_from_df(visits_df, doctors_df, nurses_df)
    medicine_forecast = predict_medicine_demand_from_df(visits_df)
    
    data = {
        "staffing": staffing_data,
        "medicine_forecast": medicine_forecast
    }
    # Cache for 1 hour
    cache.set(cache_key, data, 3600)
    return JsonResponse(data)

@login_required
def analytics_drilldown(request):
    """
    Returns detailed data for a specific chart dimension and value.
    Example: ?dimension=department&value=Cardiology
    """
    dimension = request.GET.get('dimension')
    value = request.GET.get('value')
    
    if not dimension or not value:
        return JsonResponse({"error": "Missing dimension or value"}, status=400)

    # Cache drilldown results for 30 mins
    cache_key = f"drilldown_{dimension}_{value.replace(' ', '_')}"
    cached = cache.get(cache_key)
    if cached:
        return JsonResponse(cached)

    data = {"dimension": dimension, "value": value}

    if dimension == "department":
        # Details for a specific department
        
        # 1. Top Doctors in Dept (SQL aggregate - fast)
        top_docs = Visit.objects.filter(department__name=value).values('doctor__doctor_name').annotate(count=Count('doctor')).order_by('-count')[:5]
        data["top_doctors"] = list(top_docs)
        
        # 2. Symptoms & Medicines (Recent 5000 - counted in Python for flexibility)
        recent_visits = Visit.objects.filter(department__name=value).order_by('-visit_date').values('symptoms', 'medicines')[:5000]
        
        all_symptoms = []
        all_meds = []
        for v in recent_visits:
            if v['symptoms']: all_symptoms.extend([s.strip() for s in v['symptoms'].split(',') if s.strip()])
            if v['medicines']: all_meds.extend([m.strip() for m in v['medicines'].split(',') if m.strip()])
            
        data["top_symptoms"] = dict(Counter(all_symptoms).most_common(5))
        data["top_medicines"] = dict(Counter(all_meds).most_common(5))

    elif dimension == "doctor":
        # Details for a specific doctor
        doc = Doctor.objects.filter(doctor_name=value).first()
        if not doc:
            return JsonResponse({"error": "Doctor not found"}, status=404)
        
        data["profile"] = {
            "name": doc.doctor_name,
            "dept": doc.department.name,
            "qual": doc.qualifications,
            "total_visits": Visit.objects.filter(doctor=doc).count()
        }
        
        # Top 5 Symptoms treated (Recent 5000)
        recent_symptoms = Visit.objects.filter(doctor=doc).order_by('-visit_date').values_list('symptoms', flat=True)[:5000]
        all_symptoms = []
        for v in recent_symptoms:
            if v: all_symptoms.extend([s.strip() for s in v.split(',') if s.strip()])
        data["top_symptoms"] = dict(Counter(all_symptoms).most_common(5))

    elif dimension == "case_type":
        # Trend for a specific case type (Emergency, Critical, etc)
        visits = Visit.objects.filter(case_type=value).values('visit_date')
        df = pd.DataFrame(list(visits))
        if not df.empty:
            df["visit_date"] = pd.to_datetime(df["visit_date"])
            df["ym"] = df["visit_date"].dt.strftime('%Y-%m')
            trend = df.groupby('ym').size().to_dict()
            data["trend"] = dict(sorted(trend.items())[-12:]) # Last 12 months
        else:
            data["trend"] = {}

    elif dimension == "symptom":
        # Which departments see this symptom most?
        visits = Visit.objects.filter(symptoms__icontains=value).values('department__name').annotate(count=Count('department')).order_by('-count')[:5]
        data["top_departments"] = list(visits)

    cache.set(cache_key, data, 1800)
    return JsonResponse(data)

@csrf_exempt
@login_required
def schedule_view(request):
    return render(request, "schedule.html")

@csrf_exempt
def load_schedules(request):
    schedules = []
    
    for d in Doctor.objects.all():
         schedules.append({
            "role": "Doctor",
            "name": d.doctor_name,
            "department": d.department.name,
            "shift_start": str(d.shift_start)[:5],
            "shift_end": str(d.shift_end)[:5],
            "work_days": d.work_days,
            "shift_type": d.shift_type,
            "slot_duration": d.slot_duration,
            "max_patients": d.max_patients
        })
        
    for n in Nurse.objects.all():
        schedules.append({
            "role": "Nurse",
            "name": n.nurse_name,
            "department": n.department.name,
            "shift_start": str(n.shift_start)[:5],
            "shift_end": str(n.shift_end)[:5],
            "work_days": n.work_days,
            "shift_type": n.shift_type
        })
        
    return JsonResponse(schedules, safe=False)

@csrf_exempt
def edit_schedule(request):
    role = request.POST.get("role")
    name = request.POST.get("name")
    
    try:
        if role == "Doctor":
            staff = Doctor.objects.filter(doctor_name=name).first()
            if not staff:
                raise ObjectDoesNotExist
        else:
            staff = Nurse.objects.get(nurse_name=name)
            
        staff.shift_start = request.POST.get("shift_start", staff.shift_start)
        staff.shift_end = request.POST.get("shift_end", staff.shift_end)
        staff.work_days = request.POST.get("work_days", staff.work_days)
        staff.shift_type = request.POST.get("shift_type", staff.shift_type)
        
        if role == "Doctor":
            staff.slot_duration = request.POST.get("slot_duration", staff.slot_duration)
            if request.POST.get("max_patients"):
                staff.max_patients = request.POST.get("max_patients")
                
        staff.save()
        return JsonResponse({"status": "ok"})
    except ObjectDoesNotExist:
        return JsonResponse({"status": "error", "message": "Staff not found"})


@login_required
def reception_predictions_view(request):
    """
    View for Receptionists to see predicted department loads.
    """
    cache_key = 'reception_predictions'
    cached_context = cache.get(cache_key)
    if cached_context:
        # We need to add back the non-cached context bits like 'is_reception'
        cached_context['is_reception'] = True
        return render(request, "dashboards/reception_predictions.html", cached_context)

    # 1. Fetch Visit Data - Limit to last 100k for performance
    visits = Visit.objects.all().order_by('-visit_date')[:250000].values('department__name', 'visit_date', 'created_at')
    
    context = {}
    if visits:
        df = pd.DataFrame(list(visits))
        df.rename(columns={'department__name': 'department'}, inplace=True)
        
        # 2. Predict Load for next 3 months
        predictions = predict_department_load_over_time_from_df(df, months=3)
        
        # 3. Format for Chart/Table
        today_val = date.today()
        labels = []
        for i in range(1, 4):
            future = today_val + timedelta(days=30*i) 
            labels.append(future.strftime("%b %Y"))
            
        chart_data = {
            "labels": labels,
            "datasets": []
        }
        
        colors = ['#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6', '#ec4899', '#6366f1']
        
        for idx, (dept, preds) in enumerate(predictions.items()):
            color = colors[idx % len(colors)]
            chart_data["datasets"].append({
                "label": dept,
                "data": preds,
                "borderColor": color,
                "backgroundColor": color,
                "fill": False
            })
            
        context["chart_data"] = json.dumps(chart_data)
        context["predictions_table"] = predictions
        context["months"] = labels
        context["is_reception"] = True

        # 4. Staffing Analysis
        doc_qs = Doctor.objects.values('department__name')
        nurse_qs = Nurse.objects.values('department__name')
        
        doc_df = pd.DataFrame(list(doc_qs))
        if not doc_df.empty:
            doc_df.rename(columns={'department__name': 'department'}, inplace=True)
            
        nurse_df = pd.DataFrame(list(nurse_qs))
        if not nurse_df.empty:
            nurse_df.rename(columns={'department__name': 'department'}, inplace=True)
            
        staffing_analysis = analyze_staffing_from_df(df, doc_df, nurse_df)
        context["staffing_analysis"] = staffing_analysis

        # 5. NEW: Peak Traffic Analysis (Day & Week Focus)
        traffic_data = analyze_peak_traffic_from_df(df)
        context["peak_weekly"] = json.dumps(traffic_data["weekly"])
        context["peak_daily"] = json.dumps(traffic_data["daily"])
        
    # Cache for 2 hours
    cache.set(cache_key, context, 7200)
    return render(request, "dashboards/reception_predictions.html", context)
